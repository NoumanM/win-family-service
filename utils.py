from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.service import Service
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.select import Select
from webdriver_manager.chrome import ChromeDriverManager
import time
import pandas as pd
import openpyxl
from config.config import Data
from page_xpaths import Elements
import os


BASE_DIR = os.path.dirname(os.path.realpath(__file__))

def get_chrome_driver(headless=False):
    try:
        options = webdriver.ChromeOptions()
        # options.add_argument('--incognito')
        path = f"{BASE_DIR}/cd"
        options.add_argument(f"--user-data-dir={path}")
        options.add_argument("--log-level=3")
        options.add_argument('--no-sandbox')

        if headless:
            options.add_argument('--headless')

        options.add_argument('--start-maximized')
        driver = webdriver.Chrome(service=Service(executable_path=ChromeDriverManager().install()), options=options)
        time.sleep(2)
        driver.get(Data.url)
        time.sleep(2)
        btn_click_with_action_chains(driver, Elements.next_button_on_username_page)
        send_keys_with_action_chains(driver, Elements.password_input_field, Data.password)
        btn_click_with_action_chains(driver, Elements.login_button_on_password_screen)
        return driver
    except Exception as e:
        print(e)
        print('------------------- Generation the New Driver')
        get_chrome_driver()



def send_keys_with_action_chains(driver, xpath, text, wait_time=60, previouse_clear=False):
    element = WebDriverWait(driver, wait_time).until(
        EC.visibility_of_element_located((By.XPATH, xpath)))

    action = ActionChains(driver)
    action.move_to_element(element).perform()
    if previouse_clear:
        element.clear()

    action.click().send_keys(text).perform()

def btn_click_with_action_chains(driver, xpath, wait_time=60):
    element = WebDriverWait(driver, wait_time).until(
        EC.element_to_be_clickable((By.XPATH, xpath)))

    action = ActionChains(driver)
    action.move_to_element(element).perform()
    action.click(element).perform()

def execute_script_based_click(driver, xpath, timeout=60):
    el = WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((By.XPATH, xpath)))
    driver.execute_script("arguments[0].click();", el)

def read_excel_file():
    file_path = 'myData.xlsx'
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook['Sheet1']
    max_row = sheet.max_row
    max_column = sheet.max_column
    column_names = {}
    cell_values = []
    count = 0
    for row in sheet.iter_rows(values_only=True):
        column_count = 0
        for cell in row:
            if column_count >= max_column:
                break
            column_count += 1
            if count == 0:
                column_names[cell] = ""
        count += 1
        if count == 1:
            break

    for row in sheet.iter_rows(values_only=True):
        column_count = 0
        if column_count == 0:
            column_count += 1
            continue

def get_data_from_Excel():
    file_path = 'myData.xlsx'
    df = pd.read_excel(file_path)
    data_as_list_of_dicts = df.to_dict(orient='records')
    return data_as_list_of_dicts


def select_by_text(driver, by_locator, text):
    select_element = WebDriverWait(driver, 10).until(EC.element_to_be_clickable(by_locator))
    select = Select(select_element)
    select.select_by_visible_text(text)
